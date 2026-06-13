"""
Avaliação RAGAS de pipeline RAG servido via n8n webhook.

Métricas:
  • Faithfulness     F = |V|/|S|  — resposta ancorada no contexto?
  • Context Relevancy — o contexto recuperado é relevante para a pergunta?

Uso (Fluxo em Duas Etapas):
    1. Coletar respostas:
       python avaliacao.py --step1-only
    
    2. Avaliar planilha preenchida manualmente:
       python avaliacao.py --step2-eval debug_respostas_n8n.xlsx
"""
import ast
import json
import logging
import math
import os
import re
import sys
import time
import argparse
from pathlib import Path
from typing import Optional
import openai
import requests
import pandas as pd

# ──────────────────────────────────────────────────────────────────────────────
# LOGGING
# ──────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURAÇÕES
# ──────────────────────────────────────────────────────────────────────────────
N8N_WEBHOOK_URL   = "http://localhost:5678/webhook/invoke_n8n_agent"
OPENAI_API_KEY    = os.environ.get("OPENAI_API_KEY", "")
JUDGE_LLM_MODEL   = "gpt-4o-mini"           # ou "gpt-4o" para maior precisão
JUDGE_EMBED_MODEL = "text-embedding-3-small" # ou "text-embedding-3-large"
N_QUESTIONS_AR    = 3                        # perguntas geradas para Answer Relevancy
WEBHOOK_TIMEOUT   = 120
ANSWER_KEYS  = ("output", "answer", "response", "text", "result", "message")
CONTEXT_KEYS = ("contexts", "context", "retrieved_contexts", "sources",
                "documents", "chunks", "retrievedDocs", "retrieved_docs")

DEFAULT_QUESTIONS: list[str] = [
    "Pergunta 1"
    "Pergunta 2"
]

# ──────────────────────────────────────────────────────────────────────────────
# HELPERS — OPENAI
# ──────────────────────────────────────────────────────────────────────────────
def _openai_client() -> openai.OpenAI:
    return openai.OpenAI(api_key=OPENAI_API_KEY)


def check_openai() -> None:
    if not OPENAI_API_KEY:
        log.error(
            "Variavel de ambiente OPENAI_API_KEY nao definida.\n"
            "  Execute: export OPENAI_API_KEY='sk-...'"
        )
        sys.exit(1)
    log.info("Verificando acesso a API OpenAI …")
    try:
        _openai_client().models.retrieve(JUDGE_LLM_MODEL)
    except Exception as exc:
        log.error("Nao foi possivel conectar a API OpenAI: %s", exc)
        sys.exit(1)
    log.info("OpenAI OK — modelo verificado: %s", JUDGE_LLM_MODEL)


# ──────────────────────────────────────────────────────────────────────────────
# HELPERS — CONTEXTOS
# ──────────────────────────────────────────────────────────────────────────────
def _parse_list_string(value: str) -> Optional[list[str]]:
    for parser in (json.loads, ast.literal_eval):
        try:
            parsed = parser(value)
            if isinstance(parsed, list):
                return [str(x).strip() for x in parsed if str(x).strip()]
        except Exception:
            pass
    return None


def normalize_contexts(raw) -> list[str]:
    if raw is None:
        return []
    if isinstance(raw, list):
        flat: list[str] = []
        for item in raw:
            if isinstance(item, dict):
                text = (item.get("pageContent") or item.get("page_content")
                        or item.get("text") or item.get("content") or "")
                if text.strip():
                    flat.append(text.strip())
            elif isinstance(item, list):
                flat.extend(str(x).strip() for x in item if str(x).strip())
            else:
                s = str(item).strip()
                if s:
                    flat.append(s)
        return flat
    if isinstance(raw, str):
        s = raw.strip()
        if not s:
            return []
        if s.startswith("["):
            parsed = _parse_list_string(s)
            if parsed is not None:
                return parsed
        return [s]
    return [str(raw).strip()] if str(raw).strip() else []


# ──────────────────────────────────────────────────────────────────────────────
# DIAGNÓSTICO
# ──────────────────────────────────────────────────────────────────────────────
def _warn_language(answer: str, question: str) -> None:
    thai_chars = sum(1 for c in answer if "\u0e00" <= c <= "\u0e7f")
    if thai_chars > 10:
        log.warning(
            "  IDIOMA ERRADO: Resposta em tailandes detectada para '%s...'\n"
            "    Corrija o System Prompt do AI Agent no n8n:\n"
            "    'Always respond in the same language as the user question.'",
            question[:60],
        )


def _warn_no_contexts(question: str) -> None:
    log.warning(
        "  CONTEXTOS VAZIOS para '%s...'\n"
        "    O no 'Respond to Webhook' nao esta a devolver os chunks.\n"
        "    Adicione ao Response Body do n8n:\n"
        "    \"contexts\": \"{{ $('Vector Store Tool').all().map(i => i.json.pageContent) }}\"\n"
        "    Faithfulness sera penalizado enquanto isso nao for corrigido.",
        question[:60],
    )


# ──────────────────────────────────────────────────────────────────────────────
# CONSULTA N8N
# ──────────────────────────────────────────────────────────────────────────────
def query_n8n(question: str) -> tuple[str, list[str]]:
    try:
        resp = requests.post(
            N8N_WEBHOOK_URL,
            json={"chatInput": question},
            timeout=WEBHOOK_TIMEOUT,
        )
    except requests.exceptions.Timeout:
        log.warning("Timeout ao chamar n8n para: %s", question[:80])
        return "", []
    except Exception as exc:
        log.warning("Erro de rede ao chamar n8n: %s", exc)
        return "", []

    if resp.status_code != 200:
        log.warning("n8n HTTP %s: %s", resp.status_code, resp.text[:300])
        return "", []

    raw_text = resp.text.strip()
    log.info("  Resposta bruta (%d bytes): %s", len(raw_text), raw_text[:400])

    if not raw_text:
        log.warning("  Body vazio — verifique o no 'Respond to Webhook' no n8n.")
        return "", []

    try:
        data: dict = resp.json()
    except Exception:
        log.warning("  Resposta nao e JSON — usando como texto puro.")
        return raw_text, []

    answer = ""
    for key in ANSWER_KEYS:
        val = data.get(key)
        if val and isinstance(val, str) and val.strip():
            answer = val.strip()
            log.info("  Campo answer: '%s'", key)
            break

    raw_ctx = None
    for key in CONTEXT_KEYS:
        val = data.get(key)
        if val is not None:
            raw_ctx = val
            log.info("  Campo contexts: '%s'", key)
            break

    contexts = normalize_contexts(raw_ctx)

    if answer:
        _warn_language(answer, question)

    if not contexts and answer:
        _warn_no_contexts(question)
        contexts = [""]

    return answer, contexts


# ──────────────────────────────────────────────────────────────────────────────
# COLETA
# ──────────────────────────────────────────────────────────────────────────────
def collect_responses(questions: list[str]) -> list[dict]:
    rows: list[dict] = []
    total = len(questions)
    for i, question in enumerate(questions, 1):
        log.info("[%d/%d] %s", i, total, question[:100])
        answer, contexts = query_n8n(question)
        if not answer:
            log.warning("  Resposta vazia — pergunta ignorada.")
            continue
        log.info("  answer=%d chars | contexts=%d chunk(s)", len(answer), len(contexts))
        rows.append({
            "question": question,
            "answer":   answer,
            "contexts": contexts,
            "reference": answer,
        })
        if i < total:
            time.sleep(0.5)
    return rows


# ──────────────────────────────────────────────────────────────────────────────
# MÉTRICAS RAGAS — implementação direta via OpenAI (sem LangChain/ragas)
# Referência: Es et al., 2024 — "RAGAS: Automated Evaluation of RAG"
# ──────────────────────────────────────────────────────────────────────────────

def _chat(messages: list[dict], json_mode: bool = True) -> str:
    """Chama o LLM juiz e devolve o texto da resposta."""
    kwargs: dict = dict(
        model=JUDGE_LLM_MODEL,
        messages=messages,
        temperature=0,
    )
    if json_mode:
        kwargs["response_format"] = {"type": "json_object"}
    resp = _openai_client().chat.completions.create(**kwargs)
    return resp.choices[0].message.content or ""


def _cosine(a: list[float], b: list[float]) -> float:
    dot   = sum(x * y for x, y in zip(a, b))
    na    = math.sqrt(sum(x ** 2 for x in a))
    nb    = math.sqrt(sum(x ** 2 for x in b))
    return dot / (na * nb) if na * nb > 0 else 0.0


def _get_embeddings(texts: list[str]) -> list[list[float]]:
    resp = _openai_client().embeddings.create(model=JUDGE_EMBED_MODEL, input=texts)
    return [e.embedding for e in resp.data]


# ── Faithfulness  F = |V| / |S|  ──────────────────────────────────────────────
def _compute_faithfulness(question: str, answer: str, contexts: list[str]) -> float:
    """
    1) Decompõe a resposta em statements atômicos.
    2) Verifica quais são suportados pelo contexto.
    F = supported / total
    """
    context_str = "\n".join(contexts) if contexts else ""
    if not answer.strip() or not context_str.strip():
        return float("nan")

    # Passo 1 — extração de statements
    extract_prompt = (
        "Given a question and answer, create one or more atomic statements "
        "from each sentence in the given answer.\n"
        f"question: {question}\nanswer: {answer}\n\n"
        'Return ONLY a JSON object: {"statements": ["stmt1", "stmt2", ...]}'
    )
    try:
        raw = _chat([{"role": "user", "content": extract_prompt}])
        statements: list[str] = json.loads(raw).get("statements", [])
    except Exception as exc:
        log.warning("Faithfulness — erro ao extrair statements: %s", exc)
        return float("nan")

    if not statements:
        return float("nan")

    # Passo 2 — verificação contra o contexto
    stmts_text = "\n".join(f"statement: {s}" for s in statements)
    verify_prompt = (
        "Consider the given context and the following statements, then determine "
        "whether each is supported by the information in the context.\n\n"
        f"context: {context_str}\n\n{stmts_text}\n\n"
        'Return ONLY a JSON object: {"verdicts": ["yes", "no", ...]} '
        "(one verdict per statement, same order)."
    )
    try:
        raw = _chat([{"role": "user", "content": verify_prompt}])
        verdicts: list[str] = json.loads(raw).get("verdicts", [])
    except Exception as exc:
        log.warning("Faithfulness — erro ao verificar statements: %s", exc)
        return float("nan")

    supported = sum(1 for v in verdicts if str(v).strip().lower() == "yes")
    return round(supported / len(statements), 4)


# ── Context Relevance  CR = |S_ext| / |S_total|  ──────────────────────────────
def _compute_context_relevance(question: str, contexts: list[str]) -> float:
    """
    Extrai as frases do contexto que são cruciais para responder a pergunta.
    CR = frases_extraídas / total_de_frases
    """
    context_str = "\n".join(contexts) if contexts else ""
    if not context_str.strip():
        return float("nan")

    total_sentences = len([s for s in re.split(r"(?<=[.!?])\s+", context_str.strip()) if s.strip()])
    if total_sentences == 0:
        return float("nan")

    prompt = (
        "Please extract relevant sentences from the provided context that can "
        "potentially help answer the following question. If no relevant sentences "
        'are found, return {"extracted_sentences": []}.\n\n'
        f"question: {question}\ncontext: {context_str}\n\n"
        'Return ONLY a JSON object: {"extracted_sentences": ["sent1", "sent2", ...]}'
    )
    try:
        raw = _chat([{"role": "user", "content": prompt}])
        extracted: list[str] = json.loads(raw).get("extracted_sentences", [])
    except Exception as exc:
        log.warning("Context Relevance — erro ao extrair frases: %s", exc)
        return float("nan")

    return round(len(extracted) / total_sentences, 4)


# ── Answer Relevancy  AR = mean(sim(q, q_i))  ─────────────────────────────────
def _compute_answer_relevancy(question: str, answer: str) -> float:
    """
    Gera N perguntas a partir da resposta e mede a similaridade cosseno
    com a pergunta original usando embeddings.
    AR = média das similaridades
    """
    if not answer.strip():
        return float("nan")

    generated: list[str] = []
    for _ in range(N_QUESTIONS_AR):
        prompt = (
            f"Generate a question for the given answer.\nanswer: {answer}\n\n"
            'Return ONLY a JSON object: {"question": "..."}'
        )
        try:
            raw = _chat([{"role": "user", "content": prompt}], json_mode=True)
            q = json.loads(raw).get("question", "").strip()
            if q:
                generated.append(q)
        except Exception as exc:
            log.warning("Answer Relevancy — erro ao gerar pergunta: %s", exc)

    if not generated:
        return float("nan")

    try:
        embeddings = _get_embeddings([question] + generated)
    except Exception as exc:
        log.warning("Answer Relevancy — erro ao obter embeddings: %s", exc)
        return float("nan")

    orig = embeddings[0]
    sims = [_cosine(orig, e) for e in embeddings[1:]]
    return round(sum(sims) / len(sims), 4)


# ──────────────────────────────────────────────────────────────────────────────
# AVALIAÇÃO — loop manual sem ragas/langchain
# ──────────────────────────────────────────────────────────────────────────────
def run_ragas(rows: list[dict]) -> pd.DataFrame:
    log.info("Iniciando avaliacao com OpenAI (%d amostras)...", len(rows))
    records = []

    for i, row in enumerate(rows, 1):
        question = row.get("question", "")
        answer   = row.get("answer",   "")
        contexts = row.get("contexts", [])
        log.info("[%d/%d] Avaliando: %s", i, len(rows), question[:80])

        f  = _compute_faithfulness(question, answer, contexts)
        cr = _compute_context_relevance(question, contexts)
        ar = _compute_answer_relevancy(question, answer)

        log.info(
            "  faithfulness=%.3f | context_relevancy=%.3f | answer_relevancy=%.3f",
            f if not math.isnan(f) else -1,
            cr if not math.isnan(cr) else -1,
            ar if not math.isnan(ar) else -1,
        )

        records.append({
            "question":          question,
            "answer":            answer,
            "contexts":          json.dumps(contexts, ensure_ascii=False),
            "faithfulness":      f,
            "context_relevancy": cr,
            "answer_relevancy":  ar,
        })

        if i < len(rows):
            time.sleep(0.5)   # respeitar rate limits

    df = pd.DataFrame(records)
    for col in ("faithfulness", "context_relevancy", "answer_relevancy"):
        df[col] = pd.to_numeric(df[col], errors="coerce")

    text_cols = [c for c in df.columns if c not in ("faithfulness", "context_relevancy", "answer_relevancy")]
    df[text_cols] = df[text_cols].fillna("")
    return df


# ──────────────────────────────────────────────────────────────────────────────
# RELATÓRIO EXCEL — formatação completa por coluna
# ──────────────────────────────────────────────────────────────────────────────
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers as xl_numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

_C_HEADER_DARK  = "1F3864"   
_C_HEADER_MED   = "2E75B6"   
_C_HEADER_LIGHT = "D6E4F7"   
_C_GREEN        = "C6EFCE"   
_C_GREEN_F      = "276221"
_C_YELLOW       = "FFEB9C"   
_C_YELLOW_F     = "7D6608"
_C_RED          = "FFC7CE"   
_C_RED_F        = "9C0006"
_C_ALT_ROW      = "EEF4FB"   

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=11, name="Arial") -> Font:
    return Font(bold=bold, color=color, size=size, name=name)

def _border_thin() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _score_fill(val: float) -> tuple:
    if val >= 0.7:
        return _fill(_C_GREEN),  _C_GREEN_F
    if val >= 0.4:
        return _fill(_C_YELLOW), _C_YELLOW_F
    return _fill(_C_RED), _C_RED_F

def _score_label(val: float) -> str:
    if val >= 0.7:
        return "BOM"
    if val >= 0.4:
        return "MEDIO"
    return "BAIXO"

def _set_col_width(ws, col_idx: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col_idx)].width = width

def _auto_wrap(cell) -> None:
    cell.alignment = Alignment(wrap_text=True, vertical="top")

def _write_header_row(ws, row: int, headers: list[str],
                      bg: str, fg: str = "FFFFFF", size: int = 11) -> None:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill    = _fill(bg)
        cell.font    = _font(bold=True, color=fg, size=size)
        cell.border  = _border_thin()
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

def _get_ctx_col(df):
    for c in ("context_relevancy", "nv_context_relevance", "context_relevance"):
        if c in df.columns:
            return c
    return None


def _get_ar_col(df):
    for c in ("answer_relevancy", "answer_relevance"):
        if c in df.columns:
            return c
    return None

def _sheet_resultados(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Resultados")
    ws.freeze_panes = "A2"
    ctx_col = _get_ctx_col(df)
    ar_col  = _get_ar_col(df)

    headers = [
        "#", "Pergunta",
        "Resposta do Agente",
        "Contextos Recuperados",
        "Faithfulness\n(0–1)",
        "Status\nFaithfulness",
        "Context Relevancy\n(0–1)",
        "Status\nContext Relevancy",
        "Answer Relevancy\n(0–1)",
        "Status\nAnswer Relevancy",
        "Avaliacao\nGeral",
    ]
    _write_header_row(ws, 1, headers, _C_HEADER_DARK, size=10)
    ws.row_dimensions[1].height = 36

    col_widths = [5, 45, 60, 55, 14, 18, 14, 18, 14, 18, 14]
    for i, w in enumerate(col_widths, 1):
        _set_col_width(ws, i, w)

    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        row_fill = _fill(_C_ALT_ROW) if r_idx % 2 == 0 else None

        def _to_float(v):
            try: return float(v) if str(v).strip() not in ("", "nan", "NaN") else None
            except: return None

        f_val  = _to_float(row.faithfulness)
        cr_val = _to_float(getattr(row, ctx_col, None)) if ctx_col else None
        ar_val = _to_float(getattr(row, ar_col,  None)) if ar_col  else None

        vals_nonnone = [v for v in [f_val, cr_val, ar_val] if v is not None]
        geral = round(sum(vals_nonnone) / len(vals_nonnone), 3) if vals_nonnone else None

        contexts_str = row.contexts if isinstance(row.contexts, str) else ""

        values = [
            r_idx - 1,
            row.question,
            row.answer,
            contexts_str,
            round(f_val,  3) if f_val  is not None else "N/A",
            _score_label(f_val)  if f_val  is not None else "N/A",
            round(cr_val, 3) if cr_val is not None else "N/A",
            _score_label(cr_val) if cr_val is not None else "N/A",
            round(ar_val, 3) if ar_val is not None else "N/A",
            _score_label(ar_val) if ar_val is not None else "N/A",
            geral if geral is not None else "N/A",
        ]

        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _border_thin()
            cell.font   = _font(size=10)
            _auto_wrap(cell)
            if row_fill:
                cell.fill = row_fill

        for col_offset, score_val in [(5, f_val), (7, cr_val), (9, ar_val)]:
            if score_val is not None:
                bg, fg = _score_fill(score_val)
                for delta in [0, 1]:
                    target = ws.cell(row=r_idx, column=col_offset + delta)
                    target.fill = bg
                    target.font = _font(bold=True, color=fg, size=10)

        if geral is not None:
            bg, fg = _score_fill(geral)
            ws.cell(row=r_idx, column=11).fill = bg
            ws.cell(row=r_idx, column=11).font = _font(bold=True, color=fg, size=10)

        ws.row_dimensions[r_idx].height = 60

def _sheet_resumo(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Resumo")
    ctx_col = _get_ctx_col(df)

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = "Resumo das Metricas RAGAS"
    t.fill      = _fill(_C_HEADER_DARK)
    t.font      = _font(bold=True, color="FFFFFF", size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    stats_headers = ["Metrica", "Media", "Minimo", "Maximo", "Desvio Padrao", "Interpretacao"]
    _write_header_row(ws, 2, stats_headers, _C_HEADER_MED, size=10)

    metric_labels = {
        "faithfulness":     "Faithfulness",
    }
    if ctx_col:
        metric_labels[ctx_col] = "Context Relevancy"
    ar_col = _get_ar_col(df)
    if ar_col:
        metric_labels[ar_col] = "Answer Relevancy"

    interp = {
        "faithfulness":          "Mede se a resposta e suportada pelo contexto recuperado",
        "context_relevancy":     "Mede se o contexto recuperado e relevante para a pergunta",
        "nv_context_relevance":  "Mede se o contexto recuperado e relevante para a pergunta",
        "answer_relevancy":      "Mede se a resposta gerada endereça a pergunta original",
    }

    for r_off, (col, label) in enumerate(metric_labels.items(), 3):
        vals = pd.to_numeric(df.get(col, pd.Series(dtype=float)), errors="coerce").dropna()
        if vals.empty:
            continue
        row_data = [
            label,
            round(vals.mean(), 3),
            round(vals.min(),  3),
            round(vals.max(),  3),
            round(vals.std(),  3),
            interp.get(col, ""),
        ]
        row_fill = _fill(_C_ALT_ROW) if r_off % 2 == 0 else None
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_off, column=c_idx, value=val)
            cell.border = _border_thin()
            cell.font   = _font(size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if row_fill:
                cell.fill = row_fill
            if c_idx == 2:
                try:
                    bg, fg = _score_fill(float(val))
                    cell.fill = bg
                    cell.font = _font(bold=True, color=fg, size=10)
                except Exception:
                    pass
        ws.row_dimensions[r_off].height = 32

    col_widths_resumo = [22, 10, 10, 10, 14, 70]
    for i, w in enumerate(col_widths_resumo, 1):
        _set_col_width(ws, i, w)

    ws["A6"].value = ""
    ws.merge_cells("A7:C7")
    h = ws["A7"]
    h.value     = "Scores por Pergunta"
    h.fill      = _fill(_C_HEADER_MED)
    h.font      = _font(bold=True, color="FFFFFF", size=11)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 22

    per_q_headers = ["#", "Pergunta (resumida)", "Faithfulness", "Context Relevancy", "Answer Relevancy", "Media Geral"]
    _write_header_row(ws, 8, per_q_headers, _C_HEADER_LIGHT, fg="1F3864", size=10)

    for r_off, row in enumerate(df.itertuples(index=False), 9):
        def _to_float(v):
            try: return float(v) if str(v).strip() not in ("", "nan", "NaN") else None
            except: return None

        f_val  = _to_float(row.faithfulness)
        cr_val = _to_float(getattr(row, ctx_col, None)) if ctx_col else None
        ar_val = _to_float(getattr(row, _get_ar_col(df), None)) if _get_ar_col(df) else None

        vals_nn = [v for v in [f_val, cr_val, ar_val] if v is not None]
        geral   = round(sum(vals_nn) / len(vals_nn), 3) if vals_nn else None

        row_vals = [
            r_off - 8,
            (row.question[:60] + "...") if len(row.question) > 60 else row.question,
            round(f_val,  3) if f_val  is not None else "N/A",
            round(cr_val, 3) if cr_val is not None else "N/A",
            round(ar_val, 3) if ar_val is not None else "N/A",
            geral if geral is not None else "N/A",
        ]

        row_fill = _fill(_C_ALT_ROW) if r_off % 2 == 0 else None
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r_off, column=c_idx, value=val)
            cell.border    = _border_thin()
            cell.font      = _font(size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if row_fill:
                cell.fill = row_fill
            if c_idx in (3, 4, 5, 6):
                try:
                    bg, fg = _score_fill(float(val))
                    cell.fill = bg
                    cell.font = _font(bold=True, color=fg, size=10)
                except Exception:
                    pass

        ws.row_dimensions[r_off].height = 28


def _sheet_legenda(wb) -> None:
    ws = wb.create_sheet("Legenda")

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value     = "Legenda e Guia de Interpretacao"
    t.fill      = _fill(_C_HEADER_DARK)
    t.font      = _font(bold=True, color="FFFFFF", size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    items = [
        ("SCORES",          "",                  ""),
        ("Valor",           "Status",            "Interpretacao"),
        ("0.7 — 1.0",       "BOM",               "Metrica satisfatoria — pipeline funcionando bem neste aspecto"),
        ("0.4 — 0.69",      "MEDIO",             "Atencao necessaria — ha espaco de melhoria significativo"),
        ("0.0 — 0.39",      "BAIXO",             "Problema identificado — acao corretiva necessaria"),
        ("",                "",                  ""),
        ("METRICAS",        "",                  ""),
        ("Faithfulness",    "F = |V| / |S|",     "Fracao de statements da resposta suportados pelo contexto. Score 0 = pura alucinacao."),
        ("Context Relevancy","CR = |R| / |C|",   "Grau de relevancia do contexto recuperado para a pergunta original."),
        ("Answer Relevancy", "AR = mean(sim)",    "Similaridade cosseno entre a pergunta original e perguntas geradas a partir da resposta."),
    ]

    col_widths_leg = [24, 30, 80]
    for i, w in enumerate(col_widths_leg, 1):
        _set_col_width(ws, i, w)

    r = 2
    for item in items:
        is_section = item[1] == "" and item[2] == "" and item[0] != ""
        is_header  = item[0] == "Valor"

        for c_idx, val in enumerate(item, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border    = _border_thin()
            if is_section:
                cell.fill = _fill(_C_HEADER_MED)
                cell.font = _font(bold=True, color="FFFFFF", size=10)
            elif is_header:
                cell.fill = _fill(_C_HEADER_LIGHT)
                cell.font = _font(bold=True, color=_C_HEADER_DARK, size=10)
            else:
                cell.font = _font(size=10)
                if c_idx == 2 and val in ("BOM", "MEDIO", "BAIXO"):
                    score_map = {"BOM": 1.0, "MEDIO": 0.5, "BAIXO": 0.2}
                    bg, fg = _score_fill(score_map[val])
                    cell.fill = bg
                    cell.font = _font(bold=True, color=fg, size=10)

        ws.row_dimensions[r].height = 50 if item[2] else 20
        r += 1


def save_report(df: pd.DataFrame, path: str) -> None:
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)   

        _sheet_resultados(wb, df)
        _sheet_resumo(wb, df)
        _sheet_legenda(wb)

        wb.save(path)
        log.info("Relatorio salvo em: %s  (%d abas)", path, len(wb.sheetnames))
    except PermissionError:
        log.error("Feche '%s' no Excel e rode novamente.", path)
        sys.exit(1)


# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Avalia pipeline RAG n8n com RAGAS em duas etapas."
    )
    parser.add_argument("--questions", "-q", default=None,
                        help="Arquivo .txt com uma pergunta por linha.")
    parser.add_argument("--output", "-o", default="relatorio_ragas_n8n.xlsx",
                        help="Nome do Excel de saida.")
    parser.add_argument("--debug", action="store_true",
                        help="Logs de DEBUG.")
    
    # Argumentos para o fluxo em duas etapas
    parser.add_argument("--step1-only", action="store_true",
                        help="Apenas coleta as respostas do n8n e salva na planilha de debug para voce preencher os contextos manualmente.")
    parser.add_argument("--step2-eval", type=str, default=None,
                        help="Caminho para o arquivo Excel preenchido (ex: debug_respostas_n8n.xlsx) para rodar o RAGAS direto.")
    
    return parser.parse_args()


def load_questions(path: Optional[str]) -> list[str]:
    if path is None:
        return DEFAULT_QUESTIONS
    lines = Path(path).read_text(encoding="utf-8").splitlines()
    qs = [l.strip() for l in lines if l.strip() and not l.startswith("#")]
    if not qs:
        log.error("Nenhuma pergunta em %s", path)
        sys.exit(1)
    log.info("Carregadas %d perguntas de %s", len(qs), path)
    return qs


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────
def main() -> None:
    args = parse_args()
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)

    log.info("=" * 60)
    log.info("  AVALIACAO RAGAS — FLUXO EM DUAS ETAPAS")
    log.info("=" * 60)

    check_openai()

    # ────────────────────────────────────────────────────────────────────────
    # PASSO 2: LER EXCEL E AVALIAR
    # ────────────────────────────────────────────────────────────────────────
    if args.step2_eval:
        log.info("Lendo dados do arquivo %s para avaliacao direta...", args.step2_eval)
        try:
            df_input = pd.read_excel(args.step2_eval)
            
            # --- ADICIONA ESTA LINHA AQUI ---
            df_input = df_input.dropna(subset=['question'])
            # --------------------------------
            
        except Exception as e:
            log.error("Erro ao ler planilha: %s", e)
            sys.exit(1)
            
        rows = []
        for _, r in df_input.iterrows():
            ctx_raw = r.get("contexts", "")
            
            # Tratamento inteligente do contexto colado no Excel
            if pd.isna(ctx_raw) or not str(ctx_raw).strip():
                contexts = []
            else:
                raw_str = str(ctx_raw).strip()
                # Se parecer uma lista em formato texto, tenta decodificar
                if raw_str.startswith("[") and raw_str.endswith("]"):
                    contexts = _parse_list_string(raw_str) or [raw_str]
                else:
                    # Se for um texto solto (como o chunk sujo do metadata), encapsula numa lista
                    contexts = [raw_str]
                    
           # Garante que as células vazias do Excel não quebrem o RAGAS
            q_val = r.get("question", "")
            a_val = r.get("answer", "")
            
            question_str = str(q_val) if pd.notna(q_val) else ""
            answer_str = str(a_val) if pd.notna(a_val) else ""

            rows.append({
                "question": question_str,
                "answer": answer_str,
                "contexts": contexts,
                "reference": answer_str 
            })
            
    # ────────────────────────────────────────────────────────────────────────
    # PASSO 1: CONSULTAR N8N E GUARDAR EXCEL
    # ────────────────────────────────────────────────────────────────────────
    else:
        questions = load_questions(args.questions)
        log.info("Total de perguntas: %d", len(questions))

        rows = collect_responses(questions)
        if not rows:
            log.error("Nenhuma resposta valida obtida do n8n.")
            sys.exit(1)
        log.info("%d/%d perguntas com resposta valida.", len(rows), len(questions))

        debug_path = "debug_respostas_n8n.xlsx"
        pd.DataFrame({
            "question": [r["question"] for r in rows],
            "answer":   [r["answer"]   for r in rows],
            "contexts": [json.dumps(r["contexts"], ensure_ascii=False) for r in rows],
        }).to_excel(debug_path, index=False)
        log.info("Planilha base salva em: %s", debug_path)

        if args.step1_only:
            log.info("\n" + "=" * 60)
            log.info(" PASSO 1 CONCLUIDO!")
            log.info(" 1. Abre o ficheiro '%s' no Excel.", debug_path)
            log.info(" 2. Cola os teus contextos sujos na coluna 'contexts'.")
            log.info(" 3. Guarda o ficheiro e corre o script novamente usando:")
            log.info("    python avaliacao.py --step2-eval %s", debug_path)
            log.info("=" * 60)
            sys.exit(0)

    # ────────────────────────────────────────────────────────────────────────
    # RODAR O RAGAS E EXPORTAR RESULTADOS FINAIS
    # ────────────────────────────────────────────────────────────────────────
    try:
        df_results = run_ragas(rows)
    except Exception as exc:
        log.error("Erro durante avaliacao RAGAS: %s", repr(exc))
        sys.exit(1)

    save_report(df_results, args.output)

if __name__ == "__main__":
    main()